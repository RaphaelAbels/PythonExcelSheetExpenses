import new_sheet
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Fill, Font
from openpyxl.chart import ScatterChart, Series, Reference, PieChart
font = Font(name="Arial", size=11, bold=True, italic=False, vertAlign=None, underline=None, strike=False, color="00000000")
font_sublines = Font(name="Arial", size=20, bold=True, italic=False, vertAlign="superscript", underline=None, strike=False, color="00000000")
counter = 8


def put_months_in_rows(row, start, sheet):
    for x in range(len(months) - 1):
        sheet[row + str(start + x)] = months[x]
        font_months = sheet[row + str(start + x)]
        font_months.font = font
    sheet[row + str(start + len(months) + 1)] = months[len(months) - 1]
    font_sum = sheet[row + str(start + len(months) + 1)]
    font_sum.font = font_sublines


months = ["January:", "February:", "March:", "April:", "May:", "June:",
          "July:", "August:", "September:", "October:", "November:", "December:", "Sum:"]
months_two = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
workbook = openpyxl.Workbook()
#             Daily Sheet             #


def daily_sheet(number):
    sheet_name = workbook.create_sheet(months_two[number] + "_Day")
    sheet_name["A1"] = "This sheet is for daily expenses threw out the month"
    a122 = sheet_name["A1"]
    a122.font = font
    new_sheet.daily_expenses("A", "B", "C", "E", "G", 5, 50, sheet_name)

#                Monthly Sheet            #


def monhtly_sheet(number, overview_number):
    sheet_name = workbook.create_sheet(months_two[number] + "_Month")
    topic = "TOPIC:"
    name = "NAME:"
    value = "VALUE:"
    sum = "SUM:"
    expenses = "EXPENSES:"
    sheet_name["A1"] = "This sheet is for monthly calculations to get a better overview of expenses"
    a1 = sheet_name["A1"]
    a1.font = font
    sheet_name["A3"] = "MONTHLY ASSETS:"
    a3 = sheet_name["A3"]
    a3.font = font
    sheet_name["D3"] = "=Overview!C" + str(overview_number) + "-Overview!G" + str(overview_number)
    d3 = sheet_name["D3"]
    d3.font = font
    sheet_name["A4"] = "CURRENT BALANCE:"
    a4 = sheet_name["A4"]
    a4.font = font
    sheet_name["A6"] = "EXPENSES IN TOTAL:"
    a6 = sheet_name["A6"]
    a6.font = font
    sheet_name["A7"] = "DAILY EXPENSES"
    a7 = sheet_name["A7"]
    a7.font = font
    sheet_name["D4"] = "=D3-D22-D37-I22-I37-N22-N37-S22-S37-" + months_two[number] + "_Day!F56"
    d4 = sheet_name["D4"]
    d4.font = font
    sheet_name["D6"] = "=SUM(D22,D37,I22,I37,N22,N37,S22,S37," + months_two[number] + "_Day!F56)"
    d6 = sheet_name["D6"]
    d6.font = font
    sheet_name["D7"] = "=" + months_two[number] + "_Day!F56"
    d7 = sheet_name["D7"]
    d7.font = font
    sheet_name["D8"] = "=Sum(D22,D37,I22,I37,N22,N37,S22,S37)"
    d8 = sheet_name["D8"]
    d8.font = font
    sheet_name["A8"] = "MONTHLY EXPENSES"
    a8 = sheet_name["A8"]
    a8.font = font
    new_sheet.monthly_expenses("A", "D", 10, 12, sheet_name)
    new_sheet.monthly_expenses("F", "I", 10, 12, sheet_name)
    new_sheet.monthly_expenses("K", "N", 10, 12, sheet_name)
    new_sheet.monthly_expenses("P", "S", 10, 12, sheet_name)
    new_sheet.monthly_expenses("A", "D", 25, 12, sheet_name)
    new_sheet.monthly_expenses("F", "I", 25, 12, sheet_name)
    new_sheet.monthly_expenses("K", "N", 25, 12, sheet_name)
    new_sheet.monthly_expenses("P", "S", 25, 12, sheet_name)

    # create chart                                              #

#                   Overview Sheet                    #


def expenses_overview(number, overview_number, sheet):
    sheet["K" + str(overview_number)] = "=Sum(" + months_two[number] + "_Month!N3," + months_two[number] + "_Month!I3)"


overview_sheet = workbook.active
overview_sheet.title = "Overview"
overview_sheet["A1"] = "This is the overview sheet to have a better look at the expenses and savings over the months"
a11 = overview_sheet["A1"]
a11.font = font
overview_sheet["A25"] = "Assets: Your budget for the month"
budget = overview_sheet["A25"]
budget.font = font
overview_sheet["A27"] = "Savings: The amount you want to save the month"
savings = overview_sheet["A27"]
savings.font = font
overview_sheet["A29"] = "Expenses: The amount you spent the month"
expenses = overview_sheet["A29"]
expenses.font = font
overview_sheet["A6"] = "Assets:"
assets_font = overview_sheet["A6"]
assets_font.font = font
put_months_in_rows("A", 8, overview_sheet)
overview_sheet["C22"] = new_sheet.sum_it_up("C", 8, 12)
overview_sheet["I6"] = "Expenses:"
sheet_font = overview_sheet["I6"]
sheet_font.font = font
overview_sheet["K22"] = new_sheet.sum_it_up("K", 8, 12)
put_months_in_rows("I", 8,  overview_sheet)
overview_sheet["E6"] = "Savings:"
savings = overview_sheet["E6"]
savings.font = font
put_months_in_rows("E", 8, overview_sheet)
overview_sheet["G22"] = new_sheet.sum_it_up("G", 8, 12)


#       create sheets     #

for x in range(len(months_two)):
    monhtly_sheet(x, counter)
    daily_sheet(x)
    counter = counter + 1
counter = 8
for y in range(len(months_two)):
    expenses_overview(y, counter, overview_sheet)
    counter = counter + 1


#            create charts          #
chart_sheet = workbook.create_sheet("Charts", 0)
ref_obj = openpyxl.chart.Reference(overview_sheet, min_col=11, min_row=8, max_col=11, max_row=19)
real_obj = openpyxl.chart.Series(ref_obj, title="Expenses 2020 per month")
chart_obj = openpyxl.chart.PieChart()
chart_obj.title = "Expenses in 2020 TOTAL"
chart_obj.append(real_obj)
chart_sheet.add_chart(chart_obj, "C5")



#          Introduction             #
introduction_sheet = workbook.create_sheet("Introduction (read first)", 0)
introduction_sheet["A2"] = "The following text is there for a better understanding how the sheet is working"
introduction_sheet["A4"] = "and how to have the best experience when dealing with it."


#            Save Sheet             #
workbook.save("Calculations_2020.xlsx")
