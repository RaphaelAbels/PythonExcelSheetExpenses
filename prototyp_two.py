import new_sheet
import openpyxl
from openpyxl import load_workbook


workbook = load_workbook("PROTOTYPE.xlsx")
first_sheet = workbook.active
sheet = workbook["OVERVIEW"]
sheet["D6"] = "MONTHLY ASSETS:"
# sheet["D7"] = "MONTHLY SAVINGS:"
sheet["D8"] = "MONEY LEFT FOR THIS MONTH:"
sheet["D9"] = "MONEY SPENT THIS MONTH IN TOTAL:"

sheet["D14"] = "DAILY EXPENSES IN TOTAL:"
sheet["D15"] = "MONTHLY EXPENSES IN TOTAL:"

sheet["I6"] = "ASSETS"
sheet["I7"] = "VALUE"


#       Overview Sheet         #

settings_sheet = workbook["SETTINGS"]

settings_sheet["D10"] = "NUMBER OF DIFFERENT TOPICS FOR MONTHLY EXPENSES:"
settings_sheet["D12"] = "NUMBER OF MONTHLY EXPENSES PER TOPIC:"
settings_sheet["D14"] = "APPROXIMATE NUMBER OF DAILY EXPENSES A MONTH:"

settings_sheet["K10"] = 16
settings_sheet["K12"] = 16
settings_sheet["K14"] = "COUNT"

#        Settings Sheet          #


def create_monthly_expenses(number_of_topics, number_of_rows, msheet):
    for x in range(int(number_of_topics)):
        if x % 4 == 0:
            new_sheet.monthly_expenses("A", "D", 9, number_of_rows, msheet)
            if x > 4:
                new_sheet.monthly_expenses("A", "D", 9 + number_of_rows + 1, number_of_rows, msheet)

        elif x % 4 == 1:
            new_sheet.monthly_expenses("F", "I", 9, number_of_rows, msheet)
            if x > 4:
                new_sheet.monthly_expenses("F", "I", 9 + number_of_rows + 1, number_of_rows, msheet)

        elif x % 4 == 2:
            new_sheet.monthly_expenses("K", "N", 9, number_of_rows, msheet)
            if x > 4:
                new_sheet.monthly_expenses("K", "N", 9 + number_of_rows + 1, number_of_rows, msheet)

        elif x % 4 == 3:
            new_sheet.monthly_expenses("P", "S", 9, number_of_rows, msheet)
            if x > 4:
                new_sheet.monthly_expenses("P", "S", 9 + number_of_rows + 1, number_of_rows, msheet)


monthly_sheet = workbook["MONTHLY EXPENSES"]

different_topics = int(settings_sheet.cell(10, 11).value)
row_count = int(settings_sheet.cell(12, 11).value)
create_monthly_expenses(different_topics, row_count, monthly_sheet)

#      Monthly Expenses          #
print(settings_sheet.cell(10, 11).value)

workbook.save("PROTOTYPE.xlsx")
