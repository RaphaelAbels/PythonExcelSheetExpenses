import openpyxl
import new_sheet

workbook = openpyxl.Workbook()

#          OVERVIEW            #
first_sheet = workbook.active
first_sheet.title = "SCHEDULE"
first_sheet["D6"] = "Monthly Assets"
first_sheet["D8"] = "Money left for this Month"
first_sheet["D9"] = "Money spent this Month"

#          SETTINGS              #

settings_sheet = workbook.create_sheet("SETTINGS")
settings_sheet["D10"] = "NUMBER OF DIFFERENT TOPICS FOR MONTHLY EXPENSES:"
settings_sheet["D12"] = "NUMBER OF MONTHLY EXPENSES PER TOPIC:"
settings_sheet["D14"] = "APPROXIMATE NUMBER OF DAILY EXPENSES A MONTH:"

# settings_sheet["K10"] = 4
# settings_sheet["K12"] = 6
# settings_sheet["K14"] = "COUNT"


#       MONTHLY EXPENSES         #

month_sheet = workbook.create_sheet("MONTHLY EXPENSES")


#        DAILY EXPENSES          #

daily_sheet = workbook.create_sheet("DAILY EXPENSES")


#       SAVE WORKBOOK          #


workbook.save("PROJECT SHEET.xlsx")
