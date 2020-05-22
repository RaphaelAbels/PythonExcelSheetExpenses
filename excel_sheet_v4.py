import new_sheet
import new_years
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Fill, Font
from openpyxl.chart import ScatterChart, Series, Reference, PieChart, BarChart3D
from openpyxl.worksheet.datavalidation import DataValidation
font = Font(name="Arial", size=11, bold=True, italic=False, vertAlign=None, underline=None, strike=False, color="00000000")
font_sublines = Font(name="Arial", size=20, bold=True, italic=False, vertAlign="superscript", underline=None, strike=False, color="00000000")
counter = 8

workbook = openpyxl.Workbook()
overview_sheet = workbook.active

subjects = ["Entertainment", "Subscriptions", "Groceries", "Other", "Insurances", "Car", "Rent", "Utilities"]


def daily_expenses(count, identifier, date, row, end, start, times, sheet):
    sheet[count + str(start)] = "NUMBER:"
    n1 = sheet[count + str(start)]
    n1.font = font
    sheet[identifier + str(start)] = "SUBJECT:"
    s1 = sheet[identifier + str(start)]
    s1.font = font
    sheet[date + str(start)] = "DATE:"
    n2 = sheet[date + str(start)]
    n2.font = font
    sheet[row + str(start)] = "IDENTIFIER:"
    n3 = sheet[row + str(start)]
    n3.font = font
    sheet[end + str(start)] = "Price:"
    n4 = sheet[end + str(start)]
    n4.font = font
    sheet[count + str(start + times + 1)] = "SUM:"
    n5 = sheet[count + str(start + times + 1)]
    n5.font = font
    sheet[end + str(start + times + 1)] = new_sheet.sum_it_up(end, start + 1, times)
    n6 = sheet[end + str(start + times + 1)]
    n6.font = font
    counter = 1
    for x in range(start + 1, start + times + 1):
        sheet[count + str(x)] = counter
        xy = sheet[count + str(x)]
        xy.font = font
        counter = counter + 1
    data_validation = DataValidation(type="list", formula1='"Entertainment,Subscriptions,Groceries,Other,Insurances,Car,Rent,Utilities"' , allow_blank=True)
    date_validation = DataValidation(type="date")
    sheet.add_data_validation(date_validation)
    sheet.add_data_validation(data_validation)
    data_validation.add(identifier + str(start + 1) + ":" + identifier + str(start + times))
    date_validation.add(date + str(start + 1) + ":" + date + str(start + times))


def put_subject_in_day_sheet(where, count, sheet):
    for x in range(len(subjects)):
        sheet[where + str(count + x)] = subjects[x]
        sheet_font = sheet[where + str(count + x)]
        sheet_font.font = font


def get_amount_of_a_subject(subject, row, goal, start, times, sheet):
    sum_subject = 0
    for x in range(start, start + times):
        if sheet[row + str(x)].value == subjects[subject]:
            sum_subject = int(sheet[goal + str(x)].value)

    return sum_subject;

#               Month_Sheet             #


def monthly_expenses(row, row_end, start, times, sheet, topic):
    sheet[row + str(start)] = subjects[int(topic)]
    topic_font = sheet[row + str(start)]
    topic_font.font = font_sublines
    sheet[row + str(int(start + 1))] = "Identifier"
    name_font = sheet[row + str(int(start + 1))]
    name_font.font = font
    sheet[row_end + str(start + 1)] = "Price"
    value_font = sheet[row_end + str(start + 1)]
    value_font.font = font

    for space in range(start + times, start + times + 1):
        sheet[row + str(space)] = "Sum:"
        sum_font = sheet[row + str(space)]
        sum_font.font = font_sublines
        sheet[row_end + str(space)] = new_sheet.sum_it_up(row_end, start + 2, times - 2)
        sum_sum_font = sheet[row_end + str(space)]
        sum_sum_font.font = font_sublines


def create_month_v4(month, overview_number):
    sheet_name = workbook.create_sheet(new_years.months_two[int(month)] + "_Month")
    sheet_name["A1"] = "This sheet is for payments effecting " + new_years.months_two[int(month)] + ".It is split in eight different subjects"
    a1 = sheet_name["A1"]
    a1.font = font
    sheet_name["A3"] = "Current Balance " + new_years.months_two[int(month)] + ":"
    sheet_name_font = sheet_name["A3"]
    sheet_name_font.font = font_sublines
    sheet_name["D3"] = "=Overview!C" + str(overview_number) + "-Overview!G" + str(overview_number)  + "-Overview!K" + str(overview_number)
    d3 = sheet_name["D3"]
    d3.font = font_sublines
    sheet_name["F3"] = "Daily Payments " + new_years.months_two[int(month)] + ":"
    f3 = sheet_name["F3"]
    f3.font = font_sublines
    sheet_name["I3"] = "=" + new_years.months_two[month] + "_Day!H81"
    i3 = sheet_name["I3"]
    i3.font = font_sublines
    sheet_name["K3"] = "Montly Payments " + new_years.months_two[int(month)] + ":"
    k3 = sheet_name["K3"]
    k3.font = font_sublines
    sheet_name["N3"] = "=SUM(D22,D37,I22,I37,N22,N37,S22,S37," + new_years.months_two[month] + "_Day!F56)"
    n3 = sheet_name["N3"]
    n3.font = font_sublines
    monthly_expenses("A", "D", 10, 12, sheet_name, 0)
    monthly_expenses("F", "I", 10, 12, sheet_name, 1)
    monthly_expenses("K", "N", 10, 12, sheet_name, 2)
    monthly_expenses("P", "S", 10, 12, sheet_name, 3)
    monthly_expenses("A", "D", 25, 12, sheet_name, 4)
    monthly_expenses("F", "I", 25, 12, sheet_name, 5)
    monthly_expenses("K", "N", 25, 12, sheet_name, 6)
    monthly_expenses("P", "S", 25, 12, sheet_name, 7)


#           Day_Sheet                   #

def create_day_v4(month):
    day_sheet = workbook.create_sheet(new_years.months_two[int(month)] + "_Day")
    day_sheet["A1"] = "This sheet is for payments you effect any day in " + new_years.months_two[int(month)] + "."
    daily_expenses("A", "D", "B", "F", "H", 5, 75, day_sheet)
    put_subject_in_day_sheet("K", 5, day_sheet)
    day_sheet["N5"] = get_amount_of_a_subject(0, "D", "H", 6, 74, day_sheet)


#           overview sheet              #
overview_sheet.title = "Overview"
overview_sheet["A1"] = "This is the overview sheet to have a better look at the expenses and savings over the months"
a11 = overview_sheet["A1"]
a11.font = font
overview_sheet["A25"] = "Assets: Your budget for the month"
budget = overview_sheet["A25"]
budget.font = font
overview_sheet["A27"] = "Savings: The amount you want to save the month"
savings = overview_sheet["A27"]
savings.font = font
overview_sheet["A29"] = "Expenses: The amount you spent the month"
expenses = overview_sheet["A29"]
expenses.font = font
overview_sheet["A6"] = "Assets:"
assets_font = overview_sheet["A6"]
assets_font.font = font_sublines
new_years.put_months_in_rows("A", 8, overview_sheet)
overview_sheet["C22"] = new_sheet.sum_it_up("C", 8, 12)
c22 = overview_sheet["C22"]
c22.font = font_sublines
overview_sheet["I6"] = "Expenses:"
sheet_font = overview_sheet["I6"]
sheet_font.font = font_sublines
overview_sheet["K22"] = new_sheet.sum_it_up("K", 8, 12)
k22 = overview_sheet["K22"]
k22.font = font_sublines
new_years.put_months_in_rows("I", 8,  overview_sheet)
overview_sheet["E6"] = "Savings:"
savings = overview_sheet["E6"]
savings.font = font_sublines
new_years.put_months_in_rows("E", 8, overview_sheet)
overview_sheet["G22"] = new_sheet.sum_it_up("G", 8, 12)
g22 = overview_sheet["G22"]
g22.font = font_sublines

#           create workbook             #


#       create sheets     #

for x in range(len(new_years.months_two)):
    create_month_v4(x, counter)
    create_day_v4(x)
    counter = counter + 1

counter = 8
for y in range(len(new_years.months_two)):
    new_years.expenses_overview(y, counter, overview_sheet)
    counter = counter + 1


#            create charts          #
chart_sheet = workbook.create_sheet("Charts", 0)
expenses_in_total = BarChart3D()
expenses_in_total.style = 48
ref_obj = openpyxl.chart.Reference(overview_sheet, min_col=11, min_row=8, max_col=11, max_row=19)
ref_names = Reference(overview_sheet, min_col=9, min_row=8, max_col=9, max_row=19)
real_obj = openpyxl.chart.Series(ref_obj, title="Expenses 2020 per month")
expenses_in_total.title = "Expenses in 2020 TOTAL"
expenses_in_total.add_data(data=ref_obj, titles_from_data=False)
expenses_in_total.set_categories(ref_names)
chart_sheet.add_chart(expenses_in_total, "B5")

#               Savings             #
sparing_sheet = workbook.create_sheet("Things I want to buy", 0)
sparing_sheet["A1"] = "This sheet is for potential calculations when sparing for a product."
sparing_sheet["A2"] = "Calculates how much money you have to save in order to buy your product"
sparing_sheet["D6"] = "Name:"
d6_sheet = sparing_sheet["D6"]
d6_sheet.font = font_sublines
sparing_sheet["G6"] = "Price:"
g6_sheet = sparing_sheet["G6"]
g6_sheet.font = font_sublines
sparing_sheet["J6"] = "How much I have to save for it:"
j6_sheet = sparing_sheet["J6"]
j6_sheet.font = font_sublines
sparing_sheet["O6"] = "Price I bought the product:"
m6_sheet = sparing_sheet["O6"]
m6_sheet.font = font_sublines
sparing_sheet["T6"] = "Money saved on product:"
p6_sheet = sparing_sheet["T6"]
p6_sheet.font = font_sublines


sparing_sheet["J7"] = "=G7-Overview!G22"
sparing_sheet["T7"] = "=G7-O7"

#          Introduction             #


introduction_sheet = workbook.create_sheet("Introduction (read first)", 0)
introduction_sheet["A2"] = "The following text is there for a better understanding how the sheet is working"
introduction_a2 = introduction_sheet["A2"]
introduction_a2.font = font_sublines
introduction_sheet["A4"] = "and how to have the best experience when dealing with it."
introduction_a4 = introduction_sheet["A4"]
introduction_a4.font = font_sublines
introduction_sheet["A5"] = "The following keywords describe the meaning of the different words used in the sheet."
introduction_a5 = introduction_sheet["A5"]
introduction_a5.font = font_sublines
introduction_sheet["A7"] = "Things I want to Buy:"
introduction_a7 = introduction_sheet["A7"]
introduction_a7.font = font_sublines
introduction_sheet["A9"] = "Name: The name of the product you want to buy it"
introduction_a9 = introduction_sheet["A9"]
introduction_a9.font = font_sublines
introduction_sheet["H9"] = "Price: The price of the product you saw when you wanted to buy it"
introduction_h9 = introduction_sheet["H9"]
introduction_h9.font = font_sublines
introduction_sheet["Q9"] = "How much I have to save for it: The amount that you have to save for it, including the saving over the months"
introduction_q9 = introduction_sheet["Q9"]
introduction_q9.font = font_sublines
introduction_sheet["A10"] = "Price I bought the product: The actual price at which you bought the product"
introduction_a10 = introduction_sheet["A10"]
introduction_a10.font = font_sublines
introduction_sheet["K10"] = "Money saved on product: The difference between the price you saw it and the price you bought it"
introduction_k10 = introduction_sheet["K10"]
introduction_k10.font = font_sublines
introduction_sheet["A12"] = "Charts:"
introduction_a12 = introduction_sheet["A12"]
introduction_a12.font = font_sublines
introduction_sheet["A14"] = "Expenses in 2020 TOTAL: The expenses threw out 2020 over the months"
introduction_a14 = introduction_sheet["A14"]
introduction_a14.font = font_sublines
introduction_sheet["A17"] = "Overview:"
introduction_a17 = introduction_sheet["A17"]
introduction_a17.font = font_sublines
introduction_sheet["A19"] = "Assets: The amount you want to spend over the month/year"
introduction_a19 = introduction_sheet["A19"]
introduction_a19.font = font_sublines
introduction_sheet["I19"] = "Savings: The amount you want to save this month/year (reduces expenses by that amount)"
introduction_i19 = introduction_sheet["I19"]
introduction_i19.font = font_sublines
introduction_sheet["T19"] = "Expenses: The amount you spend threw out the month/year"
introduction_t19 = introduction_sheet["T19"]
introduction_t19.font = font_sublines
introduction_sheet["A21"] = "_Month:"
introduction_a21 = introduction_sheet["A21"]
introduction_a21.font = font_sublines
introduction_sheet["A23"] = "Identifier: The name of the product you have to pay"
introduction_a23 = introduction_sheet["A23"]
introduction_a23.font = font_sublines
introduction_sheet["H23"] = "Price: The price you have to pay"
introduction_h23 = introduction_sheet["H23"]
introduction_h23.font = font_sublines
introduction_sheet["A25"] = "_Day:"
introduction_a25 = introduction_sheet["A25"]
introduction_a25.font = font_sublines
introduction_sheet["A27"] = "Date: The day you bought the product"
introduction_a27 = introduction_sheet["A27"]
introduction_a27.font = font_sublines
introduction_sheet["F27"] = "Subject: The category you bought the product from"
introduction_f27 = introduction_sheet["F27"]
introduction_f27.font = font_sublines
introduction_sheet["M27"] = "Identifier: The name of the product you have to pay"
introduction_m27 = introduction_sheet["M27"]
introduction_m27.font = font_sublines
introduction_sheet["T27"] = "Price: The price you have to pay"
introduction_t27 = introduction_sheet["T27"]
introduction_t27.font = font_sublines



#          save workbook                #


workbook.save("Calculations_2020_v4.xlsx")
